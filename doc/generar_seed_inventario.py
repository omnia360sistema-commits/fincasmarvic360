"""
generar_seed_inventario.py
Lee doc/inventaro.xlsx y genera doc/seed_inventario_historico.sql
con el catálogo de productos y los registros históricos de inventario de Marvic.

Uso:
  cd /home/pedro/Escritorio/PC/fincasmarvic-main
  python3 doc/generar_seed_inventario.py
"""

import openpyxl
import sys
import os
import urllib.request
import urllib.error
import json
from collections import defaultdict
from datetime import datetime

# ── Configuración ──────────────────────────────────────────────────────────────
EXCEL_PATH  = 'doc/inventaro.xlsx'
OUTPUT_SQL  = 'doc/seed_inventario_historico.sql'
CREATED_BY  = 'seed_historico_2026'

# Leer vars de entorno — intentar .env si no están exportadas
def load_env():
    env = {}
    for fname in ['.env', '.env.local']:
        if os.path.exists(fname):
            with open(fname) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        k, _, v = line.partition('=')
                        env[k.strip()] = v.strip().strip('"').strip("'")
    return env

_env = load_env()
SUPABASE_URL = os.getenv('VITE_SUPABASE_URL', _env.get('VITE_SUPABASE_URL', '')).rstrip('/')
SUPABASE_KEY = os.getenv('VITE_SUPABASE_ANON_KEY',
               _env.get('VITE_SUPABASE_PUBLISHABLE_KEY',
               _env.get('VITE_SUPABASE_ANON_KEY', '')))

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: No se encontraron VITE_SUPABASE_URL / VITE_SUPABASE_PUBLISHABLE_KEY en .env")
    sys.exit(1)

# ── Mapeo ubicaciones Excel → nombre exacto en Supabase ───────────────────────
LOC_MAP = {
    'BARDA':     'Cabezal Finca La Barda',
    'POLIGONO':  'Nave Polígono Finca La Barda',
    'CONCEPCION':'Nave Finca La Concepción',
    'LONSORDO':  'Nave Finca Lonsordo',
    'CIEZA':     'Nave + Cabezal Finca Collados + Cabezal Brazo Virgen',
    'COLLADOS':  'Nave + Cabezal Finca Collados + Cabezal Brazo Virgen',
}
LOC_ABONOS = 'Semillero'

# ── Reglas de clasificación por categoría (slug → palabras clave) ─────────────
# El primer match gana. Los más específicos van primero.
CAT_RULES = [
    ('maquinaria_grande',  ['MOTOR', 'BOMBA', 'DESBROZADORA', 'FLEJADORA',
                            'KARCHER', 'COMPRESOR', 'MOCHILA', 'SUBMARINO',
                            'ALARGADERA', 'RADIAL', 'CAÑON']),
    ('plastico',           ['PLASTICO', 'FILM', 'VALLA CONEJO']),
    ('manta_termica',      ['MANTA', 'HOKAIDO', 'AGROTEXTIL']),
    ('material_riego',     ['CTR ', 'CINTA', 'CODO', 'ENLACE', 'EMPALME',
                            'CASQUILLO', 'ELECTROVALVULA', 'HILAS',
                            'LATIGUILLO', 'LLAVE ', 'LLAVE A',
                            'PE 63', 'T 110', 'T 90', 'T 75',
                            'TAPONES', 'UNION ', 'VALVULA', 'RELOJ',
                            'COMPUERTA', 'ARQUETA', 'ROLLO CINTA',
                            'ROLLO GOMA', 'ROLLO PE', 'ROLLO FLEJE',
                            'REMOVEDOR', 'BIDON']),
    ('aperos_manuales',    ['AZADA', 'PALA', 'PICO', 'RASQUETA', 'CAPAZO',
                            'CUBO', 'SERRUCHO', 'TIJERA', 'CORBO',
                            'CARRETILLA', 'TRASPALETA', 'PINZA']),
    ('material_diverso',   []),   # fallback
]

def classify(nombre: str) -> str:
    n = nombre.upper()
    for slug, keywords in CAT_RULES:
        if not keywords:
            return slug          # fallback
        if any(kw in n for kw in keywords):
            return slug
    return 'material_diverso'

# ── Helpers SQL ───────────────────────────────────────────────────────────────
def sql_str(v):
    if v is None:
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

def sql_num(v):
    if v is None:
        return 'NULL'
    try:
        return str(round(float(v), 4))
    except (TypeError, ValueError):
        return 'NULL'

def sql_date(d: str) -> str:
    return f"'{d}T00:00:00+00:00'"

# ── Consultar Supabase con urllib (sin dependencias externas) ─────────────────
def supabase_get(table, select='id,nombre'):
    url = f"{SUPABASE_URL}/rest/v1/{table}?select={select}"
    req = urllib.request.Request(url, headers={
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
    })
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        print(f"ERROR HTTP {e.code} consultando {table}: {e.read().decode()}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR consultando {table}: {e}")
        sys.exit(1)

print("Consultando Supabase...")
ubicaciones_raw = supabase_get('inventario_ubicaciones', 'id,nombre')
categorias_raw  = supabase_get('inventario_categorias',  'id,nombre,slug')

ubicaciones_map = {r['nombre']: r['id'] for r in ubicaciones_raw}
categorias_map  = {r['slug']:   r['id'] for r in categorias_raw}
categorias_nombre_map = {r['slug']: r['nombre'] for r in categorias_raw}

print(f"  Ubicaciones: {list(ubicaciones_map.keys())}")
print(f"  Categorías:  {list(categorias_map.keys())}")

# Verificar mapeo
for excel_loc, supa_nombre in LOC_MAP.items():
    if supa_nombre not in ubicaciones_map:
        print(f"  ⚠ AVISO: '{supa_nombre}' (Excel: {excel_loc}) NO encontrada en Supabase")
if LOC_ABONOS not in ubicaciones_map:
    print(f"  ⚠ AVISO: '{LOC_ABONOS}' NO encontrada en Supabase")

# ── Parsear Excel ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws_campo  = wb['CAMPO']
ws_abonos = wb['ABONOS Y FITOS']

# Catálogo: (nombre_upper, slug_cat) → (precio, cat_id, unidad_defecto)
productos_catalogo = {}
# Registros históricos
registros = []
# Problemas
sin_ubicacion  = []
sin_categoria  = []

def parse_cantidad(v):
    """Convierte un valor a (float, nota_str_o_None).
    Si no es numérico devuelve (0.0, texto_original)."""
    if v is None:
        return (0.0, None)
    try:
        return (round(float(v), 4), None)
    except (TypeError, ValueError):
        return (0.0, str(v).strip())

def add_producto_y_registro(nombre_raw, finca_excel, cantidad, precio, fecha_iso):
    if nombre_raw is None:
        return
    nombre = str(nombre_raw).strip().upper()
    if nombre in ('DESCRIPCION', 'TOTALES', 'EXISTENCIAS A 28/2/26', ''):
        return
    if finca_excel is None or not isinstance(finca_excel, str):
        return

    finca = finca_excel.strip().upper()
    supa_loc = LOC_MAP.get(finca)
    if not supa_loc:
        sin_ubicacion.append((nombre, finca))
        return
    ubi_id = ubicaciones_map.get(supa_loc)
    if not ubi_id:
        sin_ubicacion.append((nombre, supa_loc))
        return

    slug_cat = classify(nombre)
    cat_id   = categorias_map.get(slug_cat)
    if not cat_id:
        sin_categoria.append((nombre, slug_cat))
        return

    cant_num, cant_nota = parse_cantidad(cantidad)

    # Guardar en catálogo (precio más reciente gana si hay duplicado)
    key = (nombre, slug_cat)
    productos_catalogo[key] = (precio, cat_id, 'unidades')

    # Registrar snapshot
    registros.append({
        'nombre':   nombre,
        'ubi_id':   ubi_id,
        'cat_id':   cat_id,
        'slug_cat': slug_cat,
        'cantidad': cant_num,
        'cantidad_nota': cant_nota,
        'precio':   precio,
        'fecha':    fecha_iso,
    })

# ── CAMPO Sección 1 (filas 1–110) ────────────────────────────────────────────
# Estructura: DESCRIPCION | STOCK_NOV30 | STOCK_DIC31 | PRECIO | IMPORTE | FINCA
print("\nParsando CAMPO Sección 1 (Nov/Dic 2025)...")
for i, row in enumerate(ws_campo.iter_rows(values_only=True), 1):
    if i > 110:
        break
    desc, s_nov, s_dic, precio, _importe, finca = row
    if desc is None or desc == 'DESCRIPCION':
        continue
    # Stock Nov-30
    if s_nov is not None:
        add_producto_y_registro(desc, finca, s_nov, precio, '2025-11-30')
    # Stock Dic-31
    if s_dic is not None:
        add_producto_y_registro(desc, finca, s_dic, precio, '2025-12-31')

# ── CAMPO Sección 2 (filas 114–229) ──────────────────────────────────────────
# Estructura: DESCRIPCION | STOCK_ENE31 | PRECIO | FINCA
print("Parsando CAMPO Sección 2 (Ene 2026)...")
for i, row in enumerate(ws_campo.iter_rows(values_only=True), 1):
    if i < 114 or i > 229:
        continue
    desc = row[0]
    s_ene, precio, finca = row[1], row[2], row[3]
    if desc is None or desc in ('DESCRIPCION', 'TOTALES'):
        continue
    if s_ene is not None:
        add_producto_y_registro(desc, finca, s_ene, precio, '2026-01-31')

# ── ABONOS Y FITOS ────────────────────────────────────────────────────────────
# Estructura (col B→G): DESCRIPCION | NOV30 | ENE07 | FEB28 | PRECIO | IMPORTE
print("Parsando ABONOS Y FITOS (3 fechas)...")
ubi_semillero = ubicaciones_map.get(LOC_ABONOS)
cat_fitos_id  = categorias_map.get('fitosanitarios_abonos')

for i, row in enumerate(ws_abonos.iter_rows(values_only=True), 1):
    if i < 7:
        continue
    desc = row[1]
    if desc is None:
        continue
    desc = str(desc).strip().upper()
    if not desc:
        continue
    nov30, ene07, feb28, precio = row[2], row[3], row[4], row[5]

    # Registrar en catálogo
    key = (desc, 'fitosanitarios_abonos')
    productos_catalogo[key] = (precio, cat_fitos_id, 'unidades')

    # Registrar snapshots (solo si hay valor y tenemos la ubicación/categoría)
    if ubi_semillero and cat_fitos_id:
        for (val, fecha) in [
            (nov30, '2025-11-30'),
            (ene07, '2026-01-07'),
            (feb28, '2026-02-28'),
        ]:
            if val is not None:
                cant_num, cant_nota = parse_cantidad(val)
                registros.append({
                    'nombre':   desc,
                    'ubi_id':   ubi_semillero,
                    'cat_id':   cat_fitos_id,
                    'slug_cat': 'fitosanitarios_abonos',
                    'cantidad': cant_num,
                    'cantidad_nota': cant_nota,
                    'precio':   precio,
                    'fecha':    fecha,
                })

# ── Generar SQL ───────────────────────────────────────────────────────────────
print(f"\nGenerando SQL...")

# Distribucion para el resumen
by_date  = defaultdict(int)
by_cat   = defaultdict(int)
for r in registros:
    by_date[r['fecha']] += 1
    by_cat[r['slug_cat']] += 1

lines = [
    "-- ================================================================",
    "-- SEED HISTÓRICO INVENTARIO MARVIC 360",
    f"-- Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    f"-- Productos catálogo: {len(productos_catalogo)}",
    f"-- Registros históricos: {len(registros)}",
    "-- EJECUTAR EN SUPABASE SQL EDITOR",
    "-- REQUISITO: Bloques 0 y 1 SQL ya ejecutados",
    "-- ================================================================",
    "",
    "BEGIN;",
    "",
    "-- ── CATÁLOGO DE PRODUCTOS ─────────────────────────────────────────",
    "-- ON CONFLICT actualiza precio si el producto ya existe",
    "",
]

for (nombre, slug_cat), (precio, cat_id, unidad_def) in sorted(productos_catalogo.items()):
    lines.append(
        f"INSERT INTO inventario_productos_catalogo "
        f"(nombre, categoria_id, precio_unitario, unidad_defecto, created_by) VALUES "
        f"({sql_str(nombre)}, '{cat_id}'::uuid, {sql_num(precio)}, "
        f"{sql_str(unidad_def)}, {sql_str(CREATED_BY)}) "
        f"ON CONFLICT (nombre, categoria_id) DO UPDATE "
        f"SET precio_unitario = EXCLUDED.precio_unitario;"
    )

lines += [
    "",
    "-- ── REGISTROS HISTÓRICOS ──────────────────────────────────────────",
    "-- Usa subquery para obtener producto_id del catálogo recién insertado",
    "",
]

for r in registros:
    # Si la cantidad original era texto, incluirla en descripcion
    nota = r.get('cantidad_nota')
    desc = r['nombre'] if nota is None else f"{r['nombre']} [stock: {nota}]"
    lines.append(
        f"INSERT INTO inventario_registros "
        f"(ubicacion_id, categoria_id, producto_id, cantidad, unidad, "
        f"descripcion, precio_unitario, created_by, created_at) "
        f"SELECT "
        f"'{r['ubi_id']}'::uuid, '{r['cat_id']}'::uuid, "
        f"ipc.id, {sql_num(r['cantidad'])}, 'unidades', "
        f"{sql_str(desc)}, {sql_num(r['precio'])}, "
        f"{sql_str(CREATED_BY)}, {sql_date(r['fecha'])} "
        f"FROM inventario_productos_catalogo ipc "
        f"WHERE ipc.nombre = {sql_str(r['nombre'])} "
        f"AND ipc.categoria_id = '{r['cat_id']}'::uuid;"
    )

lines += ["", "COMMIT;", ""]

with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

# ── Resumen ───────────────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print(f"RESUMEN GENERACIÓN SQL")
print(f"{'='*60}")
print(f"Productos únicos al catálogo : {len(productos_catalogo)}")
print(f"Registros históricos totales : {len(registros)}")
print(f"\nRegistros por fecha:")
for d, c in sorted(by_date.items()):
    print(f"  {d} : {c:>4} registros")
print(f"\nRegistros por categoría:")
for slug, c in sorted(by_cat.items(), key=lambda x: -x[1]):
    nombre_cat = categorias_nombre_map.get(slug, slug)
    print(f"  {nombre_cat:<35} : {c:>4}")

if sin_ubicacion:
    print(f"\n⚠ Sin ubicación mapeada ({len(sin_ubicacion)}):")
    for item in set(sin_ubicacion):
        print(f"  {item}")
if sin_categoria:
    print(f"\n⚠ Sin categoría mapeada ({len(sin_categoria)}):")
    for item in set(sin_categoria):
        print(f"  {item}")

print(f"\nSQL generado en: {OUTPUT_SQL}")
print(f"Líneas SQL: {len(lines)}")
print(f"\nRevisa el archivo SQL antes de ejecutarlo en Supabase.")
print(f"Comando sugerido: head -20 {OUTPUT_SQL}")
